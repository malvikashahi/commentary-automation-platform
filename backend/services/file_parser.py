# backend/services/file_parser.py

import os
import re
import openpyxl
from typing import List, Dict, Any, Optional

from backend.models.schemas import (
    AttributionData, Contributor, SectorAttribution,
    PurchaseSalesData, Transaction, PerformanceRecord
)


class FileParser:
    """Parses all source file types for the Commentary Platform."""

    # ── Attribution Parser ───────────────────────────────────
    @staticmethod
    def parse_attribution(
        file_path: str,
        account_id: str
    ) -> AttributionData:

        wb = openpyxl.load_workbook(file_path, data_only=True)

        top_contributors: List[Contributor] = []
        bottom_contributors: List[Contributor] = []
        sector_data: List[SectorAttribution] = []
        period = ""
        benchmark_name = ""
        total_portfolio_return = 0.0
        total_benchmark_return = 0.0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            name_lower = sheet_name.lower()

            # ── Top Contributors ──────────────────────────────
            if "top" in name_lower and "contrib" in name_lower:
                rows = list(ws.iter_rows(values_only=True))
                for i, row in enumerate(rows):
                    if i == 0:
                        continue
                    if row[0] is None:
                        continue
                    try:
                        top_contributors.append(Contributor(
                            rank=i,
                            security_name=str(row[0]).strip(),
                            contribution=float(row[1]) if row[1] else 0.0
                        ))
                    except Exception:
                        continue

            # ── Bottom Contributors ───────────────────────────
            elif "bottom" in name_lower or "detract" in name_lower:
                rows = list(ws.iter_rows(values_only=True))
                for i, row in enumerate(rows):
                    if i == 0:
                        continue
                    if row[0] is None:
                        continue
                    try:
                        bottom_contributors.append(Contributor(
                            rank=i,
                            security_name=str(row[0]).strip(),
                            contribution=float(row[1]) if row[1] else 0.0
                        ))
                    except Exception:
                        continue

            # ── Sector Attribution ────────────────────────────
            elif any(
                k in name_lower
                for k in ["sector", "attrib", "summary"]
            ):
                rows = list(ws.iter_rows(values_only=True))
                header_found = False

                for row in rows:
                    if row[0] is None:
                        continue
                    cell_val = str(row[0]).strip().lower()

                    if "period" in cell_val or "quarter" in cell_val:
                        period = str(row[1]) if row[1] else ""
                    if "benchmark" in cell_val:
                        benchmark_name = str(row[1]) if row[1] else ""

                    if "sector" in cell_val or "group" in cell_val:
                        header_found = True
                        continue

                    if header_found and row[0] is not None:
                        try:
                            vals = list(row)
                            if len(vals) >= 5:
                                sector_data.append(SectorAttribution(
                                    sector=str(vals[0]).strip(),
                                    portfolio_avg_wt=float(vals[1]) if vals[1] else 0.0,
                                    portfolio_return=float(vals[2]) if vals[2] else 0.0,
                                    benchmark_return=float(vals[3]) if vals[3] else 0.0,
                                    total_variance=float(vals[4]) if vals[4] else 0.0
                                ))
                                if "total" in str(vals[0]).lower():
                                    total_portfolio_return = float(vals[2]) if vals[2] else 0.0
                                    total_benchmark_return = float(vals[3]) if vals[3] else 0.0
                        except Exception:
                            continue

        # Sort contributors
        top_contributors = sorted(
            top_contributors,
            key=lambda x: x.contribution,
            reverse=True
        )[:5]

        bottom_contributors = sorted(
            bottom_contributors,
            key=lambda x: x.contribution
        )[:5]

        return AttributionData(
            account_id=account_id,
            period=period or "4Q25",
            benchmark_name=benchmark_name or "Russell 1000 Growth Index (Total Return)",
            top_contributors=top_contributors,
            bottom_contributors=bottom_contributors,
            sector_attribution=sector_data,
            total_portfolio_return=total_portfolio_return,
            total_benchmark_return=total_benchmark_return,
            total_active_return=round(
                total_portfolio_return - total_benchmark_return, 4
            )
        )

    # ── Purchase & Sales Parser ──────────────────────────────
    @staticmethod
    def parse_purchase_sales(
        file_path: str,
        account_id: str
    ) -> PurchaseSalesData:

        wb = openpyxl.load_workbook(file_path, data_only=True)
        purchases: List[Transaction] = []
        sales: List[Transaction] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            name_lower = sheet_name.lower()
            rows = list(ws.iter_rows(values_only=True))

            is_purchase = (
                "purch" in name_lower or "buy" in name_lower
            )
            is_sale = (
                "sale" in name_lower or
                "sell" in name_lower or
                "sold" in name_lower
            )

            # Auto-detect from header row
            if not is_purchase and not is_sale:
                if rows and rows[0]:
                    h = str(rows[0][0]).lower()
                    is_purchase = "purch" in h
                    is_sale = "sale" in h or "sold" in h

            for i, row in enumerate(rows):
                if i == 0:
                    continue
                if row[0] is None:
                    continue
                try:
                    # Find rationale — last non-null string column
                    rationale = ""
                    for cell in reversed(row):
                        if (
                            cell and
                            isinstance(cell, str) and
                            len(str(cell)) > 10
                        ):
                            rationale = str(cell).strip()
                            break

                    txn = Transaction(
                        security_name=str(row[0]).strip(),
                        transaction_date=str(row[1]) if row[1] else None,
                        quantity=(
                            float(row[2])
                            if row[2] and
                            str(row[2]).replace('.', '').isdigit()
                            else None
                        ),
                        rationale=rationale or "No rationale provided."
                    )

                    if is_purchase:
                        purchases.append(txn)
                    elif is_sale:
                        sales.append(txn)
                    else:
                        purchases.append(txn)

                except Exception:
                    continue

        return PurchaseSalesData(
            account_id=account_id,
            purchases=purchases,
            sales=sales
        )

    # ── Performance Parser ───────────────────────────────────
    @staticmethod
    def parse_performance(file_path: str) -> List[PerformanceRecord]:

        wb = openpyxl.load_workbook(file_path, data_only=True)
        records: List[PerformanceRecord] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            headers: List[str] = []
            header_row = None

            for i, row in enumerate(rows):
                if row[0] is None:
                    continue
                cell = str(row[0]).lower().strip()

                # Detect header row
                if any(k in cell for k in ["account", "code", "name"]):
                    header_row = i
                    headers = [
                        str(h).lower() if h else ""
                        for h in row
                    ]
                    continue

                if header_row is not None:
                    try:
                        def get_col(keywords):
                            for kw in keywords:
                                for j, h in enumerate(headers):
                                    if kw in h and j < len(row):
                                        return row[j]
                            return None

                        account_code = (
                            get_col(["account", "code"]) or row[0]
                        )
                        period = (
                            get_col(["period", "quarter", "date"])
                            or "4Q25"
                        )
                        port_ret = (
                            get_col(["portfolio", "port", "return"])
                            or 0.0
                        )
                        bench_ret = (
                            get_col(["bench", "index"]) or 0.0
                        )
                        active = (
                            get_col(["active", "excess", "vs"])
                            or 0.0
                        )

                        if account_code:
                            records.append(PerformanceRecord(
                                account_code=str(account_code).strip(),
                                period=str(period),
                                portfolio_return=float(port_ret) if port_ret else 0.0,
                                benchmark_return=float(bench_ret) if bench_ret else 0.0,
                                active_return=float(active) if active else 0.0
                            ))
                    except Exception:
                        continue

        return records

    # ── Lead Commentary Parser ───────────────────────────────
    @staticmethod
    def parse_lead_commentary(file_path: str) -> Dict[str, Any]:
        """
        Parse Word document (.doc or .docx).
        Returns sections, placeholders, paragraphs, full_text.
        """
        ext = os.path.splitext(file_path)[1].lower()

        # ── Handle .doc files ─────────────────────────────────
        if ext == ".doc":
            try:
                import docx2txt
                full_text = docx2txt.process(file_path) or ""

                sections: Dict[str, str] = {}
                paragraphs = []
                current_section = "Introduction"
                current_content: List[str] = []
                placeholders = set()

                SECTION_KEYWORDS = [
                    "performance", "attribution", "transaction",
                    "activity", "portfolio", "overview", "market",
                    "outlook", "strategy", "summary", "positioning"
                ]

                for line in full_text.split('\n'):
                    text = line.strip()
                    if not text:
                        continue

                    tokens = re.findall(r'\[([^\]]+)\]', text)
                    for token in tokens:
                        placeholders.add(token)

                    is_header = (
                        len(text) < 80 and
                        any(kw in text.lower() for kw in SECTION_KEYWORDS) and
                        not text.endswith('.')
                    )

                    if is_header:
                        if current_content:
                            sections[current_section] = '\n'.join(
                                current_content
                            )
                        current_section = text
                        current_content = []
                    else:
                        current_content.append(text)

                    paragraphs.append({
                        "text": text,
                        "style": "Normal",
                        "is_header": is_header,
                        "bold": False
                    })

                if current_content:
                    sections[current_section] = '\n'.join(current_content)

                return {
                    "full_text": full_text,
                    "sections": sections,
                    "placeholders": list(placeholders),
                    "paragraphs": paragraphs,
                    "section_count": len(sections)
                }

            except Exception as e:
                print(f"Warning: Could not parse .doc file: {e}")
                return {
                    "full_text": "",
                    "sections": {
                        "Introduction": "Lead commentary could not be parsed."
                    },
                    "placeholders": [],
                    "paragraphs": [],
                    "section_count": 0
                }

        # ── Handle .docx files ────────────────────────────────
        else:
            try:
                from docx import Document

                doc = Document(file_path)
                sections: Dict[str, str] = {}
                current_section = "Introduction"
                current_content: List[str] = []
                full_paragraphs = []
                placeholders = set()

                SECTION_KEYWORDS = [
                    "performance", "attribution", "transaction",
                    "activity", "portfolio", "overview", "market",
                    "outlook", "strategy", "summary", "positioning"
                ]

                for para in doc.paragraphs:
                    text = para.text.strip()
                    if not text:
                        continue

                    tokens = re.findall(r'\[([^\]]+)\]', text)
                    for token in tokens:
                        placeholders.add(token)

                    is_header = (
                        para.style.name.startswith('Heading') or
                        (
                            para.runs and
                            all(
                                run.bold
                                for run in para.runs
                                if run.text.strip()
                            )
                        ) or
                        (
                            any(
                                kw in text.lower()
                                for kw in SECTION_KEYWORDS
                            ) and len(text) < 80
                        )
                    )

                    if is_header:
                        if current_content:
                            sections[current_section] = '\n'.join(
                                current_content
                            )
                        current_section = text
                        current_content = []
                    else:
                        current_content.append(text)

                    full_paragraphs.append({
                        "text": text,
                        "style": para.style.name,
                        "is_header": is_header,
                        "bold": any(
                            run.bold
                            for run in para.runs
                            if run.text.strip()
                        )
                    })

                if current_content:
                    sections[current_section] = '\n'.join(current_content)

                return {
                    "full_text": '\n'.join(
                        p['text'] for p in full_paragraphs
                    ),
                    "sections": sections,
                    "placeholders": list(placeholders),
                    "paragraphs": full_paragraphs,
                    "section_count": len(sections)
                }

            except Exception as e:
                print(f"Warning: Could not parse .docx file: {e}")
                return {
                    "full_text": "",
                    "sections": {
                        "Introduction": "Lead commentary could not be parsed."
                    },
                    "placeholders": [],
                    "paragraphs": [],
                    "section_count": 0
                }